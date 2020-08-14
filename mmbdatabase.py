# Модуль для работы с базой данных

import openpyxl  # Модуль для работы с excel файлами
from openpyxl.styles import Font, Border, Side, Alignment  # Для стилей в excel
import datetime as dt  # Модуль для получения даты и времени


class Database():
    """Класс базы данных"""

    def __init__(self, file_name):
        self.file_name = file_name  # Имя файла с базой данных
        self.file = openpyxl.load_workbook(filename=file_name)  # Объект таблицы
        self.main_sheet = self.file['main']  # Еxcel лист с основной таблицей

    def fill_ceil(self, r, c, val):
        """
        Форматированное заполнение ячейки (r, c) значением val
        r - номер строки
        с - номер колонки
        val - значение, которое нужно записать
        """
        # На всякий случай обернул в try, не знаю, нужно ли
        try:
            self.main_sheet.cell(row=r, column=c).value = val
            ceil = self.main_sheet.cell(row=r, column=c)
            ceil.font = Font(size=14)  # Разммер шрифта 14
            thin = Side(border_style="thin", color="000000")  # Тонкая рамка
            ceil.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            # Выравнивание текста по левому краю, и по центру по вертикали
            ceil.alignment = Alignment(horizontal="left", vertical="center")
        except PermissionError:
            pass

    def add_user(self, q_users, message):
        """Добавление инфы о пользователе, написавшем боту /start"""
        self.main_sheet['F1'] = q_users  # Записываем обновлённое кол-во
        self.fill_ceil(2 + q_users, 1, q_users)  # Порядковый номер
        self.fill_ceil(2 + q_users, 2, message.from_user.id)
        self.fill_ceil(2 + q_users, 3, message.from_user.first_name)
        self.fill_ceil(2 + q_users, 4, message.from_user.last_name)
        self.fill_ceil(2 + q_users, 5, message.from_user.username)
        self.fill_ceil(2 + q_users, 6, dt.datetime.now().date())
        self.fill_ceil(2 + q_users, 7, dt.datetime.now().time())

    def new_user(self, message):
        """Обработка нового пользователя"""
        quantity_of_users = int(
            self.main_sheet['F1'].value)  # Кол-во пользователей из базы
        users_id_list = [self.main_sheet.cell(row=i, column=2).value
                         for i in range(3, 3 + quantity_of_users)]
        if message.chat.id not in users_id_list:  # Если пользователя нет в базе
            quantity_of_users += 1
            self.add_user(quantity_of_users, message)

        # Если файл параллельно открыт, возникает ошибка при сохранении
        try:
            self.file.save(self.file_name)
        except PermissionError:
            print(
                'Файл с базой данных параллельно открыт в другом приложении. ' +
                'Данные не сохранены!')

    def clear_main_sheet(self):
        """
        Очистка основной таблицы базы данных:
        удаление id, имени, ника, времени отправки "/start" для каждого пользователя
        обнуление кол-ва пользователей
        """
        quantity_of_users = int(self.main_sheet["F1"].value)
        for row in range(quantity_of_users):
            for column in range(1, 8):
                self.fill_ceil(3 + row, column, '')
        self.main_sheet["F1"] = 0  # Обнуляем кол-во пользователей

        # Если файл параллельно открыт, возникает ошибка при сохранении
        try:
            self.file.save(self.file_name)
        except PermissionError:
            print(
                "Файл с базой данных параллельно открыт в другом приложении. " +
                "Изменения не сохранены!")
        else:
            print("Основной лист базы данных успешно очишен")
