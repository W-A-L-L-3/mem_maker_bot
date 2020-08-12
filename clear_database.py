# Файл для очистки различных разделов базы данных

"""
Список возможностей:
    1) Очистка основной таблицы базы данных - clear_main_database_sheet()
    2) Удаление всех папок с файлами пользователя - clear_all_users_files()
"""

import openpyxl  # Модуль для работы с excel файлами
from openpyxl.styles import Font, Border, Side, Alignment  # Для стилей в excel
# Модули для работы с файловой системой
import os
import shutil


def fill_ceil(sheet, r, c, val):
    """
    Форматированное заполнение ячейки (r, c) значением val
    r - номер строки
    с - номер колонки
    val - значение, которое нужно записать
    """
    # На всякий случай обернул в try, не знаю, нужно ли
    try:
        sheet.cell(row=r, column=c).value = val
        ceil = sheet.cell(row=r, column=c)
        ceil.font = Font(size=14)  # Разммер шрифта 14
        thin = Side(border_style="thin", color="000000")  # Тонкая черная рамка
        ceil.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # Выравнивание текста по левому краю, и по центру по вертикали
        ceil.alignment = Alignment(horizontal="left", vertical="center")
    except PermissionError:
        pass


def clear_main_database_sheet():
    """
    Очистка основной таблицы базы данных:
    удаление id, имени, ника, времени отправки "/start" для каждого пользователя
    обнуление кол-ва пользователей
    """
    global database_file_name, db
    database_main_sheet = db["main"]  # Еxcel лист с основной таблицей
    quantity_of_users = int(database_main_sheet["F1"].value)
    for row in range(quantity_of_users):
        for column in range(1, 8):
            fill_ceil(database_main_sheet, 3 + row, column, '')
    database_main_sheet["F1"] = 0  # Обнуляем кол-во пользователей

    # Если файл параллельно открыт, возникает ошибка при сохранении
    try:
        db.save(database_file_name)
    except PermissionError:
        print("Файл с базой данных параллельно открыт в другом приложении. " +
              "Изменения не сохранены!")
    else:
        print("Основной лист базы данных успешно очишен")


def clear_all_users_files(del_data=False):
    """
    Удаление всех папок с файлами пользователя
    del_data - удалять ли файл data каждого пользователя (True/False)
    """

    def check(name):
        """Проеварка на то, является ли объект с данным путём папкой"""
        return os.path.isdir(f"./users/{name}")

    e = 0  # Счётчик ошибок
    # Обход всех папок в каталоге "users"
    for folder_name in filter(check, os.listdir(path="./users")):
        try:
            if del_data:  # Удаление папки целиком (вместе с "data")
                shutil.rmtree(f"./users/{folder_name}")
            else:  # Удаление только папки "img"
                shutil.rmtree(f"./users/{folder_name}/img")
        except OSError as e:  # Если возникла ошибка при поиске данного пути
            print(f"Error: {e.filename} - {e.strerror}.")
            e += 1
        else:
            print(f"Папка {folder_name} удалена")

    if e == 0:  # Если ошибок не было
        print()
        print("Все папки успешно удалены")
    else:
        print(f"Ошибки при удалении {e} папки(ок)")


database_file_name = "database.xlsx"  # Имя файла с базой данных
db = openpyxl.load_workbook(filename=database_file_name)  # Файл с базой данных

clear_main_database_sheet()
# clear_all_users_files()
